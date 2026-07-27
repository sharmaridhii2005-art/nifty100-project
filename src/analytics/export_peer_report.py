import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class PeerReportExporter:

    def __init__(self, db_path="db/nifty100.db"):
        self.conn = sqlite3.connect(db_path)

    def export(self):

        # -----------------------------
        # Load peer percentile table
        # -----------------------------
        percentiles = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.conn
        )

        # -----------------------------
        # Load peer groups
        # -----------------------------
        peer_groups = pd.read_sql(
            """
            SELECT
                company_id,
                peer_group_name,
                is_benchmark
            FROM peer_groups
            """,
            self.conn
        )

        # -----------------------------
        # Load company names
        # (id contains company symbol)
        # -----------------------------
        companies = pd.read_sql(
            """
            SELECT
                id,
                company_name
            FROM companies
            """,
            self.conn
        )

        companies.rename(
            columns={"id": "company_id"},
            inplace=True
        )

        # -----------------------------
        # Merge company names
        # -----------------------------
        df = percentiles.merge(
            companies,
            on="company_id",
            how="left"
        )

        # -----------------------------
        # Merge benchmark flag
        # -----------------------------
        df = df.merge(
            peer_groups[
                [
                    "company_id",
                    "is_benchmark"
                ]
            ],
            on="company_id",
            how="left"
        )

        # -----------------------------
        # Create output folder
        # -----------------------------
        import os
        os.makedirs("output", exist_ok=True)

        output_file = "output/peer_comparison.xlsx"

        # -----------------------------
        # Export each peer group
        # -----------------------------
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            for group in sorted(df["peer_group_name"].dropna().unique()):

                sheet = df[
                    df["peer_group_name"] == group
                ].copy()

                wide = sheet.pivot_table(
                    index=[
                        "company_id",
                        "company_name",
                        "is_benchmark"
                    ],
                    columns="metric",
                    values="percentile_rank"
                ).reset_index()

                wide.to_excel(
                    writer,
                    sheet_name=group[:31],
                    index=False
                )

        # -----------------------------
        # Apply formatting
        # -----------------------------
        wb = load_workbook(output_file)

        green = PatternFill(
            start_color="90EE90",
            end_color="90EE90",
            fill_type="solid"
        )

        yellow = PatternFill(
            start_color="FFFF99",
            end_color="FFFF99",
            fill_type="solid"
        )

        red = PatternFill(
            start_color="FF9999",
            end_color="FF9999",
            fill_type="solid"
        )

        gold = PatternFill(
            start_color="FFD700",
            end_color="FFD700",
            fill_type="solid"
        )

        for ws in wb.worksheets:

            headers = [cell.value for cell in ws[1]]

            benchmark_col = headers.index("is_benchmark") + 1

            for row in range(2, ws.max_row + 1):

                if ws.cell(row, benchmark_col).value == 1:

                    for cell in ws[row]:
                        cell.fill = gold

                for col in range(4, ws.max_column + 1):

                    value = ws.cell(row, col).value

                    if value is None:
                        continue

                    if value >= 75:
                        ws.cell(row, col).fill = green

                    elif value <= 25:
                        ws.cell(row, col).fill = red

                    else:
                        ws.cell(row, col).fill = yellow

        wb.save(output_file)

        print("\nPeer comparison report exported successfully.")
        print(f"Saved to: {output_file}")